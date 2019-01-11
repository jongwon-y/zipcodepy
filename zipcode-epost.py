import re
import openpyxl # pip install openpyxl
from selenium import webdriver # pip install selenium
from bs4 import BeautifulSoup # pip install bs4

driver = webdriver.Chrome('./chromedriver/chromedriver') # chromedriver 경로
driver.implicitly_wait(3) # 웹페이지 최대 대기 시간 (초)

xls = openpyxl.load_workbook('sample.xlsx') # 작업할 파일
actXls = xls['Sheet1'] # 작업할 시트

result = []

for r in actXls.rows:
    adress = r[0].value
    piece = adress.split()
    result.append(piece)

rowNum = 0

for i in result:
    count = 0

    for word in i:
        p = re.compile(r'[로길]$') # 끝이 "로, 길" 로 끝나는 단어 조각 찾기
        m = p.search(word)
        if m:
            toSearch = word + ' ' + i[count+1]
            driver.get('https://www.epost.go.kr/search.RetrieveIntegrationNewZipCdList.comm') # 우편번호 찾는 사이트
            driver.find_element_by_id('keyword').send_keys(toSearch)
            driver.find_element_by_id('btnZipSearch').click()
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            zipCodeRaw = soup.select('.table_list > tbody > tr.title2 > th') # 우편번호 출력 위치
            if zipCodeRaw:
                zipCode = zipCodeRaw[0].text.strip() # 첫 번째 우편번호 사용
                actXls.cell(row=rowNum+1, column=2).value = zipCode # 우편번호를 입력할 행,열 지정
            else:
                zipCode ='can not find' # 찾을 수 없을때
                actXls.cell(row=rowNum+1, column=2).value = zipCode # 우편번호를 입력할 행,열 지정

        count = count + 1
    rowNum = rowNum + 1

xls.save(filename='new.xlsx') # 저장할 파일명
xls.close()