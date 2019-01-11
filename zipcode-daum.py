import openpyxl # pip install openpyxl
from selenium import webdriver # pip install selenium
from bs4 import BeautifulSoup # pip install bs4

driver = webdriver.Chrome('./chromedriver/chromedriver') # chromedriver 경로
driver.implicitly_wait(3) # 웹페이지 최대 대기 시간 (초)

xls = openpyxl.load_workbook('sample.xlsx') # 작업할 파일
actXls = xls['Sheet1'] # 작업할 시트

addList = []

for r in actXls.rows:
    adress = r[0].value
    addList.append(adress)

rowNum = 0

for addr in addList:
    driver.get('http://postcode.map.daum.net/search?region_name=&cq=&cpage=1&origin=http%3A%2F%2F&isp=N&isgr=N&isgj=N&ongr=&ongj=&regionid=&regionname=&roadcode=&roadname=&banner=on&indaum=off&vt=popup&am=on&ani=off&mode=transmit&sd=on&hmb=off&heb=off&asea=off&smh=off&zo=off&theme=&bit=&sit=&sgit=&sbit=&pit=&mit=&lcit=&plrg=&plrgt=1.5&us=on&msi=10&ahs=off&whas=500&zn=Y&sm=on&CWinWidth=1280&sptype=&sporgq=&fullpath=%2FC%3A%2Fworkspace%2Fdaum.html&a51=off')
    driver.find_element_by_id('region_name').send_keys(addr)
    driver.find_element_by_class_name('btn_search').click()
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    zipCodeRaw = soup.select('.txt_postcode')
    if zipCodeRaw:
        zipCode = zipCodeRaw[0].text.strip()
        actXls.cell(row=rowNum+1, column=2).value = zipCode
    else:
        actXls.cell(row=rowNum+1, column=2).value = 'can not find'
    rowNum = rowNum + 1
xls.save(filename='new.xlsx') # 저장할 파일명
xls.close()